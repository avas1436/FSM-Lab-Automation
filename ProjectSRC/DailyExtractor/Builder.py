import datetime
from abc import ABC
from typing import Any

import pandas as pd  # type: ignore
from openpyxl import load_workbook  # type: ignore

from ProjectSRC.DailyExtractor.Product import LabResult  # type: ignore
from ProjectSRC.Logger.logger_config import logger


# Interface
class ExcelAdapter(ABC):
    """یک کلاس پایه که همه ادپتر های زیر مجموعه باید از آن ارث ببرند"""

    def get_records(self):
        raise NotImplementedError


# Adapter with openpyxl
class Openpyxl(ExcelAdapter):
    """لود کردن دیتا در رم به صورت یکجا"""

    def __init__(self, file_path: str, start: int, end: int):
        self.file_path = file_path
        self.start = start
        self.end = end

    def get_records(self):
        wb = load_workbook(filename=self.file_path, data_only=True, read_only=True)
        try:
            sheet_names = wb.sheetnames
            active_sheets = sheet_names[self.start - 1 : self.end]
            for sheet in active_sheets:
                ws = wb[sheet]
                cell_range = ws["A4":"L31"]
                yield [[cell.value for cell in ls] for ls in cell_range]
        finally:
            wb.close()


# Adapter with pandas
class Pandas(ExcelAdapter):
    def __init__(self, file_path: str, start: int, end: int):
        self.file_path = file_path
        self.start = start
        self.end = end

    def get_records(self):
        # خواندن کل فایل اکسل
        xls = pd.ExcelFile(self.file_path)
        sheet_names = xls.sheet_names
        active_sheets = sheet_names[self.start - 1 : self.end]

        for sheet in active_sheets:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet,
                header=None,
                skiprows=3,
                nrows=28,
                usecols="A:L",
            )
            sheet_data = df.values.tolist()
            yield sheet_data


# Facade
class ExcelAdapterFacade:
    def __init__(
        self, file_path: str, start: int = 1, end: int = 31, engine: str = "openpyxl"
    ):
        if engine == "openpyxl":
            self.adapter = Openpyxl(file_path=file_path, start=start, end=end)
        elif engine == "pandas":
            self.adapter = Pandas(file_path=file_path, start=start, end=end)  # type: ignore
        else:
            raise ValueError("Engine must be 'pandas' or 'openpyxl'")

    def get_records(self):
        return self.adapter.get_records()


class LabResultBuilder:
    def __init__(self, raw_data: list[list]):
        self.raw_data = raw_data
        self._records: list[dict[str, Any]] = []
        # logger.debug(
        #     f"LabResultBuilder initialized with {len(raw_data)} rows of raw data"
        # )

    def parse(self):
        logger.info("Starting parse process...")

        try:
            date = self.raw_data[0][4]  # '1404/07/02'
            year, month, day = str(date).split("/")
            logger.debug(f"Extracted date: year={year}, month={month}, day={day}")
        except Exception as e:
            logger.error(f"Failed to extract date from raw_data[0][4]: {e}")
            raise

        for i, ls in enumerate(self.raw_data):
            if i < 8:
                # logger.debug(f"Skipping row {i} (header or irrelevant)")
                continue

            if isinstance(ls[0], (datetime.time, datetime.datetime)):
                time = ls[0].strftime("%H%M")
                # logger.debug(f"Row {i}: Parsed time {time}")
            else:
                logger.warning(f"Row {i} skipped: invalid time format ({type(ls[0])})")
                continue

            record = {
                "year": year,
                "month": month,
                "day": day,
                "time": time,
                "klin1": str(ls[1]) if ls[1] is not None else "Not Tested",
                "klin2": str(ls[3]) if ls[3] is not None else "Not Tested",
                "above40": (str(ls[5]) if ls[5] is not None else "Not Tested"),
                "par05": (
                    str(self.raw_data[i + 1][8])
                    if 7 < i < 20 and self.raw_data[i + 1][8] is not None
                    else "Not Tested"
                ),
                "par51": (
                    str(self.raw_data[i + 1][9])
                    if 7 < i < 20 and self.raw_data[i + 1][9] is not None
                    else "Not Tested"
                ),
                "par60": str(ls[11]) if ls[11] is not None else "Not Tested",
            }

            fields = ["klin1", "klin2", "above40", "par05"]
            if all(record[f] == "Not Tested" for f in fields):
                # logger.info(f"Row {i} skipped: all key fields are 'Not Tested'")
                continue

            # logger.debug(f"Row {i}: Record added {record}")
            self._records.append(record)

        logger.info(f"Parsing complete. {len(self._records)} records built.")
        return self

    def build(self) -> list["LabResult"]:
        # logger.info(f"Building LabResult objects from {len(self._records)} records")
        results = [LabResult(**rec) for rec in self._records]
        logger.debug(f"Successfully built {len(results)} LabResult objects")
        return results


# openpyxl adaptor test:

# excel = Openpyxl(
# file_path=r"C:\Users\abAsz\Documents\GIT\FSM Lab Automation\ProjectSRC\DailyExtractor\daily.xlsx",
#     start=1,
#     end=4,
# )

# data = excel.get_records()

# for i in range(4):
#     data_parser = LabResultBuilder(next(data))
#     lab_result = data_parser.parse().build()
#     for res in lab_result:
#         print(res.model_dump())


# pandas adaptor test:

# excel = Pandas(
#     file_path=r"F:\گزارش\1404\9. آذر\daily.xlsx",
#     start=1,
#     end=2,
# )

# data = excel.get_records()

# for i in range(2):
#     data_parser = LabResultBuilder(next(data))
#     lab_result = data_parser.parse().build()
#     for res in lab_result:
#         print(res.model_dump())


# test facade
# excel = ExcelAdapterFacade(
#     file_path=r"daily.xlsx",
#     start=1,
#     end=1,
#     engine="openpyxl",
# )

# full_data = list(excel.get_records())
# for res in full_data:
#     data_parser = LabResultBuilder(res)
#     lab_result = data_parser.parse().build()
#     for res in lab_result:
#         print(res.model_dump())
